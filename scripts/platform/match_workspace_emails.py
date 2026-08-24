"""Match the Google Workspace user export to members and write the loader's e-mail CSV.

    uv run poe match-emails \
        [--xlsx data/workspace/User_Download_15052026_173847_File.xlsx] \
        [--out data/derived/workspace-emails.csv] \
        [--min-confidence 0.90]

Workspace is the source of truth for who can sign in (ADR 0001): one cdtm.com mailbox binds
to one Member by e-mail. The roster and the scrape carry no mailbox, so this script joins
the export to ``members`` by name and writes ``slug,email,method,confidence``. The loader
(``load_community.py --emails``) then binds without making any decision of its own
(ADR 0004); low-confidence pairs go to a review file instead of the loader input.

Matching, in order, first hit wins:
  1. ``exact``      normalised "first last" equals the member's name or roster name
  2. ``local-part`` the e-mail local part (``first.last``, ``flast``, ``firstl``) equals the
                    member's slug or the same pattern built from the member's name
  3. ``fuzzy``      difflib ratio between normalised full names >= ``--min-confidence``,
                    only when it is the single best candidate by a clear margin
Anything ambiguous (two members for one mailbox, or one member for two mailboxes) is written
to the review file, never to the loader input.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from _bootstrap import ensure_repo_on_path

ensure_repo_on_path()

from sqlalchemy import text  # noqa: E402

from infrastructure.db import get_sync_engine  # noqa: E402

_NON_ALNUM = re.compile(r"[^a-z0-9 ]+")


def norm(s: str | None) -> str:
    """Lower-case ASCII letters and digits only: 'Köken, Adil' -> 'koken adil'."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower().replace("-", " ").replace("_", " ").replace(".", " ")
    s = _NON_ALNUM.sub(" ", s)
    return " ".join(s.split())


def name_variants(first: str, last: str) -> set[str]:
    f, last_n = norm(first), norm(last)
    out = {f"{f} {last_n}".strip(), f"{last_n} {f}".strip()}
    # "Anna Maria Muster" in Workspace vs "Anna Muster" on the roster: try first token only.
    if " " in f:
        out.add(f"{f.split()[0]} {last_n}")
    if " " in last_n:
        out.add(f"{f} {last_n.split()[-1]}")
    return {v for v in out if v}


def local_part_variants(first: str, last: str) -> set[str]:
    f, last_n = norm(first).replace(" ", ""), norm(last).replace(" ", "")
    if not f or not last_n:
        return set()
    return {f"{f}{last_n}", f"{f[0]}{last_n}", f"{f}{last_n[0]}", f"{f} {last_n}".replace(" ", "")}


@dataclass
class MemberRow:
    id: str
    slug: str
    name: str | None
    first_name: str | None
    last_name: str | None
    roster_name: str | None
    email: str | None

    def name_keys(self) -> set[str]:
        out: set[str] = set()
        for n in (self.name, self.roster_name):
            if n:
                out.add(norm(n))
                parts = norm(n).split()
                if len(parts) >= 2:
                    out.add(f"{parts[-1]} {' '.join(parts[:-1])}")
        if self.first_name or self.last_name:
            out |= name_variants(self.first_name or "", self.last_name or "")
        return {k for k in out if k}

    def local_keys(self) -> set[str]:
        out = {norm(self.slug).replace(" ", "")}
        if self.first_name and self.last_name:
            out |= local_part_variants(self.first_name, self.last_name)
        elif self.name and " " in self.name:
            parts = self.name.split()
            out |= local_part_variants(parts[0], parts[-1])
        return {k for k in out if k}


@dataclass
class WorkspaceRow:
    first: str
    last: str
    email: str
    status: str


def read_workspace(xlsx: Path) -> list[WorkspaceRow]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - dependency group not installed
        sys.exit("openpyxl missing: run `uv sync --group data`")
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h or "").strip().lower() for h in next(rows)]

    def col(prefix: str) -> int:
        for i, h in enumerate(header):
            if h.startswith(prefix):
                return i
        sys.exit(f"column starting with {prefix!r} not found in {xlsx.name}: {header}")

    fi, li, ei, si = col("first name"), col("last name"), col("email address"), col("status")
    out = []
    for r in rows:
        if not r or not r[ei]:
            continue
        out.append(
            WorkspaceRow(
                first=str(r[fi] or "").strip(),
                last=str(r[li] or "").strip(),
                email=str(r[ei]).strip().lower(),
                status=str(r[si] or "").strip(),
            )
        )
    return out


def load_members() -> list[MemberRow]:
    with get_sync_engine().connect() as conn:
        rows = conn.execute(
            text(
                "select id::text, slug, name, first_name, last_name, roster_name, email "
                "from members"
            )
        ).all()
    return [MemberRow(*r) for r in rows]


def match(
    members: list[MemberRow], people: list[WorkspaceRow], *, min_confidence: float
) -> tuple[list[dict], list[dict], list[dict]]:
    by_name: dict[str, list[MemberRow]] = defaultdict(list)
    by_local: dict[str, list[MemberRow]] = defaultdict(list)
    for m in members:
        for k in m.name_keys():
            by_name[k].append(m)
        for k in m.local_keys():
            by_local[k].append(m)
    all_keys = {k: ms for k, ms in by_name.items() if len(ms) == 1}

    matched: list[dict] = []
    review: list[dict] = []
    unmatched: list[dict] = []
    claimed: dict[str, str] = {}  # member id -> email

    def emit(m: MemberRow, p: WorkspaceRow, method: str, confidence: float) -> None:
        rec = {
            "slug": m.slug,
            "email": p.email,
            "method": method,
            "confidence": f"{confidence:.2f}",
            "workspace_name": f"{p.first} {p.last}".strip(),
            "member_name": m.name or m.roster_name or "",
            "status": p.status,
        }
        if m.id in claimed and claimed[m.id] != p.email:
            rec["reason"] = f"member already matched to {claimed[m.id]}"
            review.append(rec)
            return
        if m.email and m.email.lower() != p.email:
            rec["reason"] = f"member already has e-mail {m.email}"
            review.append(rec)
            return
        if confidence < min_confidence:
            rec["reason"] = "below min confidence"
            review.append(rec)
            return
        claimed[m.id] = p.email
        matched.append(rec)

    for p in people:
        # 1. exact name
        hits: list[MemberRow] = []
        for v in name_variants(p.first, p.last):
            hits.extend(by_name.get(v, []))
        uniq = {m.id: m for m in hits}
        if len(uniq) == 1:
            emit(next(iter(uniq.values())), p, "exact", 1.0)
            continue
        if len(uniq) > 1:
            for m in uniq.values():
                review.append(
                    {
                        "slug": m.slug,
                        "email": p.email,
                        "method": "exact",
                        "confidence": "0.50",
                        "workspace_name": f"{p.first} {p.last}",
                        "member_name": m.name or "",
                        "status": p.status,
                        "reason": "several members share this name",
                    }
                )
            continue
        # 2. local part of the address
        local = p.email.split("@", 1)[0]
        local_n = norm(local).replace(" ", "")
        lhits = {m.id: m for m in by_local.get(local_n, [])}
        if len(lhits) == 1:
            emit(next(iter(lhits.values())), p, "local-part", 0.95)
            continue
        # 3. fuzzy on the full name
        target = f"{norm(p.first)} {norm(p.last)}".strip()
        if target:
            scored = sorted(
                ((SequenceMatcher(None, target, k).ratio(), k) for k in all_keys),
                reverse=True,
            )[:2]
            if scored:
                best, key = scored[0]
                second = scored[1][0] if len(scored) > 1 else 0.0
                if best >= min_confidence and best - second >= 0.05:
                    emit(all_keys[key][0], p, "fuzzy", best)
                    continue
                if best >= 0.8:
                    m = all_keys[key][0]
                    review.append(
                        {
                            "slug": m.slug,
                            "email": p.email,
                            "method": "fuzzy",
                            "confidence": f"{best:.2f}",
                            "workspace_name": f"{p.first} {p.last}",
                            "member_name": m.name or "",
                            "status": p.status,
                            "reason": "fuzzy candidate, check by hand",
                        }
                    )
                    continue
        unmatched.append(
            {"email": p.email, "workspace_name": f"{p.first} {p.last}", "status": p.status}
        )
    return matched, review, unmatched


def write_csv(path: Path, rows: list[dict], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawTextHelpFormatter)
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="Workspace export; defaults to the newest *.xlsx in data/workspace/",
    )
    ap.add_argument("--out", type=Path, default=Path("data/derived/workspace-emails.csv"))
    ap.add_argument("--min-confidence", type=float, default=0.90)
    args = ap.parse_args()

    xlsx = args.xlsx
    if xlsx is None:
        candidates = sorted(Path("data/workspace").glob("*.xlsx"))
        if not candidates:
            sys.exit("no data/workspace/*.xlsx found")
        xlsx = candidates[-1]

    people = read_workspace(xlsx)
    members = load_members()
    matched, review, unmatched = match(members, people, min_confidence=args.min_confidence)

    out_dir = args.out.parent
    write_csv(
        args.out,
        matched,
        ["slug", "email", "method", "confidence", "workspace_name", "member_name", "status"],
    )
    write_csv(
        out_dir / "workspace-review.csv",
        review,
        [
            "slug",
            "email",
            "method",
            "confidence",
            "workspace_name",
            "member_name",
            "status",
            "reason",
        ],
    )
    write_csv(out_dir / "workspace-unmatched.csv", unmatched, ["email", "workspace_name", "status"])

    by_method: dict[str, int] = defaultdict(int)
    for r in matched:
        by_method[r["method"]] += 1
    print(f"workspace rows: {len(people)}  members: {len(members)}")
    print(f"matched: {len(matched)}  {dict(by_method)}")
    print(f"review: {len(review)}  unmatched: {len(unmatched)}")
    print(f"wrote {args.out}, workspace-review.csv, workspace-unmatched.csv")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
